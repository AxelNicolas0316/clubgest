import io
import os
import requests
from datetime import datetime

from flask import send_file, make_response

# Librerías para generar archivos PDF con ReportLab
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm, cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable, Image as RLImage
)
from reportlab.platypus.flowables import HRFlowable
from reportlab.pdfgen import canvas as pdfcanvas

# Librerías para generar archivos Excel con OpenPyXL
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side,
    GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, Reference


# Colores institucionales de Don Bosco para usar en los reportes
DB_NAVY      = colors.HexColor('#0a1628')
DB_BLUE      = colors.HexColor('#0d2a5e')
DB_BLUE_MID  = colors.HexColor('#163a7a')
DB_GOLD      = colors.HexColor('#d4a017')
DB_GOLD2     = colors.HexColor('#f0c040')
DB_RED       = colors.HexColor('#c0272d')
DB_WHITE     = colors.white
DB_LIGHT     = colors.HexColor('#e8f0fb')
DB_MUTED     = colors.HexColor('#7a9cc4')
DB_GRAY_ROW  = colors.HexColor('#f4f7fc')
DB_GRAY_ROW2 = colors.HexColor('#edf2f9')

LOGO_URL = 'https://donboscolatola.edu.ec/wp-content/uploads/2024/01/cropped-LOGO-DON-BOSCO.png'
LOGO_LOCAL = os.path.join(os.path.dirname(__file__), 'imagenes', 'don_bosco.png')


def _get_logo_path():
    """Devuelve path del logo: primero local, luego descarga."""
    if os.path.exists(LOGO_LOCAL):
        return LOGO_LOCAL
    try:
        resp = requests.get(LOGO_URL, timeout=5)
        if resp.status_code == 200:
            tmp = '/tmp/db_logo_tmp.png'
            with open(tmp, 'wb') as f:
                f.write(resp.content)
            return tmp
    except Exception:
        pass
    return None


def _col_headers(tipo):
    """Devuelve encabezados y keys según el tipo de informe."""
    base = {
        'nivel': {
            'headers': ['#', 'Nombres y Apellidos', 'Correo Institucional', 'Género', 'Nivel', 'Club', 'Tutor', 'Especialidad'],
            'keys':    ['_num', '_fullname', 'correo_institucional', 'genero', 'nombre_nivel', 'nombre_club', 'tutor', 'nombre_especialidad'],
        },
        'club': {
            'headers': ['#', 'Nombres y Apellidos', 'Correo Institucional', 'Género', 'Nivel', 'Club', 'Tutor', 'Especialidad'],
            'keys':    ['_num', '_fullname', 'correo_institucional', 'genero', 'nombre_nivel', 'nombre_club', 'tutor', 'nombre_especialidad'],
        },
        'especialidad': {
            'headers': ['#', 'Nombres y Apellidos', 'Correo Institucional', 'Género', 'Nivel', 'Especialidad', 'Club', 'Tutor'],
            'keys':    ['_num', '_fullname', 'correo_institucional', 'genero', 'nombre_nivel', 'nombre_especialidad', 'nombre_club', 'tutor'],
        },
    }
    return base.get(tipo, base['nivel'])


# Clase personalizada para crear un PDF con encabezado y pie de página institucionales
class DBDocTemplate(SimpleDocTemplate):
    """Template personalizado con header/footer institucional."""

    def __init__(self, buffer, titulo_corto, **kwargs):
        self.titulo_corto = titulo_corto
        super().__init__(buffer, **kwargs)

    def handle_pageBegin(self):
        self._handle_pageBegin()

    def afterPage(self):
        pass

    def _draw_bg_header(self, canvas_obj, doc):
        """Dibuja el encabezado institucional en cada página."""
        w, h = doc.pagesize
        canvas_obj.saveState()

        # Franja superior azul marino
        canvas_obj.setFillColor(DB_NAVY)
        canvas_obj.rect(0, h - 28*mm, w, 28*mm, fill=1, stroke=0)

        # Línea dorada debajo del header
        canvas_obj.setFillColor(DB_GOLD)
        canvas_obj.rect(0, h - 29.5*mm, w, 1.5*mm, fill=1, stroke=0)

        # Acento rojo izquierdo
        canvas_obj.setFillColor(DB_RED)
        canvas_obj.rect(0, h - 28*mm, 5*mm, 28*mm, fill=1, stroke=0)

        # Acento dorado en la parte roja (decorativo)
        canvas_obj.setFillColor(DB_GOLD)
        canvas_obj.rect(5*mm, h - 28*mm, 2*mm, 28*mm, fill=1, stroke=0)

        # Logo
        logo_path = _get_logo_path()
        if logo_path:
            try:
                canvas_obj.drawImage(logo_path, 10*mm, h - 26*mm, width=18*mm, height=18*mm,
                                     preserveAspectRatio=True, mask='auto')
            except Exception:
                pass

        # Textos del header
        canvas_obj.setFillColor(DB_WHITE)
        canvas_obj.setFont('Helvetica-Bold', 11)
        canvas_obj.drawString(32*mm, h - 10*mm, 'UNIDAD EDUCATIVA FISCOMISIONAL SALESIANA')
        canvas_obj.setFont('Helvetica', 9)
        canvas_obj.setFillColor(DB_GOLD2)
        canvas_obj.drawString(32*mm, h - 16*mm, 'Don Bosco La Tola · Quito, Ecuador')
        canvas_obj.setFillColor(DB_MUTED)
        canvas_obj.setFont('Helvetica', 7.5)
        canvas_obj.drawString(32*mm, h - 21*mm, 'Sistema de Gestión de Clubes Escolares — ClubGest v2.0')

        # Fecha/hora arriba derecha
        canvas_obj.setFillColor(colors.HexColor('#7a9cc4'))
        canvas_obj.setFont('Helvetica', 7)
        now_str = datetime.now().strftime('%d/%m/%Y  %H:%M')
        canvas_obj.drawRightString(w - 10*mm, h - 10*mm, f'Generado: {now_str}')
        canvas_obj.drawRightString(w - 10*mm, h - 16*mm, f'Página {doc.page}')

        # Footer
        canvas_obj.setFillColor(DB_NAVY)
        canvas_obj.rect(0, 0, w, 10*mm, fill=1, stroke=0)
        canvas_obj.setFillColor(DB_GOLD)
        canvas_obj.rect(0, 10*mm, w, 1*mm, fill=1, stroke=0)

        canvas_obj.setFillColor(DB_MUTED)
        canvas_obj.setFont('Helvetica', 6.5)
        canvas_obj.drawString(10*mm, 4*mm, 'Don Bosco E5–O6 y Los Ríos, Barrio La Tola · Quito, Ecuador · Tel: (593) 2582-493')
        canvas_obj.setFillColor(DB_GOLD2)
        canvas_obj.drawRightString(w - 10*mm, 4*mm, 'donboscolatola.edu.ec')

        canvas_obj.restoreState()

    def build(self, flowables, onFirstPage=None, onLaterPages=None, canvasmaker=None):
        def on_page(c, doc):
            self._draw_bg_header(c, doc)
        super().build(flowables, onFirstPage=on_page, onLaterPages=on_page)


def generar_pdf(datos, titulo, tipo):
    """
    Genera un PDF institucional impactante.
    
    Args:
        datos: lista de dicts con los estudiantes
        titulo: string del título del informe
        tipo: 'nivel' | 'club' | 'especialidad'
    
    Returns:
        Flask Response con el PDF
    """
    buf = io.BytesIO()
    PAGE = landscape(A4)
    w, h = PAGE

    doc = DBDocTemplate(
        buf, titulo,
        pagesize=PAGE,
        topMargin=35*mm, bottomMargin=18*mm,
        leftMargin=12*mm, rightMargin=12*mm,
        title=f'Informe ClubGest — {titulo}',
        author='Don Bosco La Tola',
        subject='Inscripciones Clubes Escolares'
    )

    styles = getSampleStyleSheet()
    story  = []

    # Título principal del PDF con información de la institución
    style_titulo = ParagraphStyle(
        'titulo', fontName='Helvetica-Bold', fontSize=16,
        textColor=DB_NAVY, alignment=TA_CENTER, spaceAfter=2*mm,
        leading=20
    )
    style_subtitulo = ParagraphStyle(
        'sub', fontName='Helvetica', fontSize=10,
        textColor=DB_BLUE_MID, alignment=TA_CENTER, spaceAfter=1*mm
    )
    style_meta = ParagraphStyle(
        'meta', fontName='Helvetica', fontSize=8,
        textColor=DB_MUTED, alignment=TA_CENTER, spaceAfter=4*mm
    )

    story.append(Spacer(1, 4*mm))
    story.append(Paragraph(f'INFORME DE INSCRIPCIONES', style_titulo))
    story.append(Paragraph(titulo, style_subtitulo))
    story.append(Paragraph(
        f'Periodo: {datetime.now().strftime("%Y")} · Total de registros: <b>{len(datos)}</b>',
        style_meta
    ))
    story.append(HRFlowable(width='100%', thickness=2, color=DB_GOLD, spaceAfter=4*mm))

    # Mostrar estadísticas rápidas como conteos de estudiantes
    if datos:
        # Contar por club
        clubs_count = {}
        for r in datos:
            c = r.get('nombre_club', '—')
            clubs_count[c] = clubs_count.get(c, 0) + 1

        # Contar por género
        gen_m = sum(1 for r in datos if 'm' in str(r.get('genero', '')).lower())
        gen_f = sum(1 for r in datos if 'f' in str(r.get('genero', '')).lower())

        style_stats_title = ParagraphStyle('st', fontName='Helvetica-Bold', fontSize=8.5,
                                            textColor=DB_BLUE, spaceAfter=2*mm)
        story.append(Paragraph('RESUMEN ESTADÍSTICO', style_stats_title))

        stat_data = [
            ['Total Inscritos', 'Masculino', 'Femenino', 'Clubes participantes', 'Registro generado'],
            [
                str(len(datos)),
                str(gen_m),
                str(gen_f),
                str(len(clubs_count)),
                datetime.now().strftime('%d/%m/%Y %H:%M')
            ]
        ]

        stat_table = Table(stat_data, colWidths=[None]*5)
        stat_table.setStyle(TableStyle([
            ('BACKGROUND',   (0,0), (-1,0), DB_BLUE),
            ('TEXTCOLOR',    (0,0), (-1,0), DB_GOLD2),
            ('FONTNAME',     (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',     (0,0), (-1,0), 7.5),
            ('FONTNAME',     (0,1), (-1,1), 'Helvetica-Bold'),
            ('FONTSIZE',     (0,1), (-1,1), 13),
            ('TEXTCOLOR',    (0,1), (-1,1), DB_NAVY),
            ('BACKGROUND',   (0,1), (-1,1), DB_GOLD2),
            ('ALIGN',        (0,0), (-1,-1), 'CENTER'),
            ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0,0), (-1,-1), [DB_BLUE, DB_GOLD2]),
            ('TOPPADDING',   (0,0), (-1,-1), 4),
            ('BOTTOMPADDING',(0,0), (-1,-1), 4),
            ('BOX',          (0,0), (-1,-1), 0.5, DB_GOLD),
            ('INNERGRID',    (0,0), (-1,-1), 0.3, colors.HexColor('#c0a830')),
        ]))
        story.append(stat_table)
        story.append(Spacer(1, 5*mm))

    # Tabla principal con la lista de todos los estudiantes inscritos
    if not datos:
        no_data = ParagraphStyle('nd', fontName='Helvetica', fontSize=11,
                                  textColor=DB_MUTED, alignment=TA_CENTER)
        story.append(Spacer(1, 20*mm))
        story.append(Paragraph('No se encontraron registros para los filtros seleccionados.', no_data))
    else:
        col_info = _col_headers(tipo)
        headers  = col_info['headers']
        keys     = col_info['keys']

        # Construir filas
        table_data = [headers]
        for idx, row in enumerate(datos, 1):
            fila = []
            for k in keys:
                if k == '_num':
                    fila.append(str(idx))
                elif k == '_fullname':
                    nombre = f"{row.get('nombres','')} {row.get('apellidos','')}".strip()
                    fila.append(nombre)
                else:
                    fila.append(str(row.get(k, '—') or '—'))
            table_data.append(fila)

        # Anchos de columna dinámicos
        total_w = w - 24*mm
        num_cols = len(headers)
        # #=8mm, nombres=55mm, correo=55mm, resto=automático
        fixed = {0: 8*mm, 1: 55*mm, 2: 55*mm}
        remaining = total_w - sum(fixed.values())
        auto_w = remaining / max(num_cols - len(fixed), 1)
        col_widths = [fixed.get(i, auto_w) for i in range(num_cols)]

        main_table = Table(table_data, colWidths=col_widths, repeatRows=1)

        # Alternating row colors
        row_styles = [
            ('BACKGROUND',    (0,0),  (-1,0),  DB_NAVY),
            ('TEXTCOLOR',     (0,0),  (-1,0),  DB_GOLD2),
            ('FONTNAME',      (0,0),  (-1,0),  'Helvetica-Bold'),
            ('FONTSIZE',      (0,0),  (-1,0),  7.5),
            ('ALIGN',         (0,0),  (-1,0),  'CENTER'),
            ('FONTNAME',      (0,1),  (-1,-1), 'Helvetica'),
            ('FONTSIZE',      (0,1),  (-1,-1), 7.5),
            ('ALIGN',         (0,0),  (0,-1),  'CENTER'),
            ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
            ('TOPPADDING',    (0,0),  (-1,-1), 4),
            ('BOTTOMPADDING', (0,0),  (-1,-1), 4),
            ('LEFTPADDING',   (0,0),  (-1,-1), 4),
            ('RIGHTPADDING',  (0,0),  (-1,-1), 4),
            ('BOX',           (0,0),  (-1,-1), 0.5, DB_GOLD),
            ('LINEBELOW',     (0,0),  (-1,0),  1, DB_GOLD),
            ('INNERGRID',     (0,1),  (-1,-1), 0.3, colors.HexColor('#d0d8e8')),
            # Club column gold
            ('TEXTCOLOR',     (5,1),  (5,-1),  DB_BLUE),
            ('FONTNAME',      (5,1),  (5,-1),  'Helvetica-Bold'),
            # Num column
            ('TEXTCOLOR',     (0,1),  (0,-1),  DB_MUTED),
            ('FONTNAME',      (0,1),  (0,-1),  'Helvetica-Bold'),
        ]
        # Alternating rows
        for i in range(1, len(table_data)):
            bg = DB_GRAY_ROW if i % 2 == 0 else DB_WHITE
            row_styles.append(('BACKGROUND', (0,i), (-1,i), bg))

        main_table.setStyle(TableStyle(row_styles))
        story.append(main_table)

    doc.build(story)
    buf.seek(0)

    filename = f'informe_{tipo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    return send_file(
        buf,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )


# Función para generar reportes profesionales en formato Excel
def generar_excel(datos, titulo, tipo):
    """
    Genera un archivo Excel profesional con diseño institucional.
    
    Args:
        datos: lista de dicts
        titulo: string del título
        tipo: 'nivel' | 'club' | 'especialidad'
    
    Returns:
        Flask Response con el .xlsx
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inscripciones'

    # Convertir valores hexadecimales de colores para usarlos en Excel
    C_NAVY   = '0a1628'
    C_BLUE   = '0d2a5e'
    C_GOLD   = 'd4a017'
    C_GOLD2  = 'f0c040'
    C_RED    = 'c0272d'
    C_WHITE  = 'FFFFFF'
    C_LIGHT1 = 'EDF4FF'
    C_LIGHT2 = 'F8FBFF'
    C_MUTED  = '7a9cc4'

    thin  = Side(style='thin',   color=C_NAVY)
    thick = Side(style='medium', color=C_GOLD)
    def border(t=False, b=False, l=False, r=False):
        return Border(
            top=thick if t else thin if t is None else Side(style=None),
            bottom=thick if b else thin if b is None else Side(style=None),
            left=thick if l else thin if l is None else Side(style=None),
            right=thick if r else thin if r is None else Side(style=None),
        )

    col_info = _col_headers(tipo)
    headers  = col_info['headers']
    keys     = col_info['keys']
    num_cols = len(headers)

    # Establecer el ancho de las columnas de la tabla
    col_w = {1: 6, 2: 35, 3: 38, 4: 12, 5: 22, 6: 28, 7: 22, 8: 22}
    for i in range(1, num_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = col_w.get(i, 20)

    # Primera fila: banda de color azul marino de la institución
    ws.row_dimensions[1].height = 14
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=c, value='')
        cell.fill = PatternFill('solid', fgColor=C_RED if c == 1 else C_NAVY)

    # Filas 2-5: información del encabezado con logo y datos de la institución
    ws.merge_cells(start_row=2, start_column=1, end_row=5, end_column=num_cols)
    header_cell = ws.cell(row=2, column=1)
    header_cell.value = 'UNIDAD EDUCATIVA FISCOMISIONAL SALESIANA\nDon Bosco La Tola'
    header_cell.font  = Font(name='Calibri', size=18, bold=True, color=C_GOLD2)
    header_cell.fill  = PatternFill('solid', fgColor=C_NAVY)
    header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for r in range(2, 6):
        ws.row_dimensions[r].height = 16
        for c in range(2, num_cols + 1):
            ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=C_NAVY)

    # Fila 6: línea de separación con color dorado
    ws.row_dimensions[6].height = 4
    for c in range(1, num_cols + 1):
        ws.cell(row=6, column=c).fill = PatternFill('solid', fgColor=C_GOLD)

    # Filas 7-8: título del informe
    ws.merge_cells(start_row=7, start_column=1, end_row=8, end_column=num_cols)
    tit = ws.cell(row=7, column=1)
    tit.value = titulo.upper()
    tit.font  = Font(name='Calibri', size=13, bold=True, color=C_NAVY)
    tit.fill  = PatternFill('solid', fgColor='E8F0FF')
    tit.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[7].height = 24
    ws.row_dimensions[8].height = 14
    for c in range(2, num_cols + 1):
        ws.cell(row=8, column=c).fill = PatternFill('solid', fgColor='E8F0FF')

    # Fila 9: información sobre cuándo se generó el informe
    ws.merge_cells(start_row=9, start_column=1, end_row=9, end_column=num_cols // 2)
    ws.merge_cells(start_row=9, start_column=num_cols // 2 + 1, end_row=9, end_column=num_cols)
    meta_l = ws.cell(row=9, column=1)
    meta_l.value = f'Total de registros: {len(datos)}'
    meta_l.font  = Font(name='Calibri', size=9, color='1a3a6a', bold=True)
    meta_l.fill  = PatternFill('solid', fgColor='D6E4F7')
    meta_l.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    meta_r = ws.cell(row=9, column=num_cols // 2 + 1)
    meta_r.value = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    meta_r.font  = Font(name='Calibri', size=9, color='1a3a6a')
    meta_r.fill  = PatternFill('solid', fgColor='D6E4F7')
    meta_r.alignment = Alignment(horizontal='right', vertical='center', indent=1)
    ws.row_dimensions[9].height = 16

    for c in range(2, num_cols + 1):
        if c != num_cols // 2 + 1:
            ws.cell(row=9, column=c).fill = PatternFill('solid', fgColor='D6E4F7')

    # Espacio en blanco para separación visual
    ws.row_dimensions[10].height = 6

    # Fila 11: nombres de las columnas de la tabla
    ROW_HEADER = 11
    ws.row_dimensions[ROW_HEADER].height = 28
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=ROW_HEADER, column=i, value=h.upper())
        c.font      = Font(name='Calibri', size=9, bold=True, color=C_GOLD2)
        c.fill      = PatternFill('solid', fgColor=C_NAVY)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = Border(
            bottom=Side(style='medium', color=C_GOLD),
            top=Side(style='thin', color=C_GOLD),
            left=Side(style='thin', color='1e3a6a'),
            right=Side(style='thin', color='1e3a6a'),
        )

    # Agregar los datos de estudiantes en las filas
    for idx, row in enumerate(datos, 1):
        excel_row = ROW_HEADER + idx
        ws.row_dimensions[excel_row].height = 18

        is_even = idx % 2 == 0
        bg_col  = C_LIGHT1 if is_even else C_LIGHT2

        for ci, k in enumerate(keys, 1):
            if k == '_num':
                val = idx
            elif k == '_fullname':
                val = f"{row.get('nombres','')} {row.get('apellidos','')}".strip()
            else:
                val = row.get(k, '—') or '—'

            c = ws.cell(row=excel_row, column=ci, value=val)
            c.fill      = PatternFill('solid', fgColor=bg_col)
            c.alignment = Alignment(vertical='center',
                                     horizontal='center' if ci == 1 else 'left',
                                     indent=0 if ci == 1 else 1)
            c.border = Border(
                bottom=Side(style='thin', color='C8D8EE'),
                left=Side(style='thin',   color='D8E8F8'),
                right=Side(style='thin',  color='D8E8F8'),
            )
            # Estilos especiales
            if ci == 1:
                c.font = Font(name='Calibri', size=9, bold=True, color=C_MUTED)
            elif ci == 2:
                c.font = Font(name='Calibri', size=9, bold=True, color=C_NAVY)
            elif k == 'nombre_club':
                c.font = Font(name='Calibri', size=9, bold=True, color=C_BLUE)
            else:
                c.font = Font(name='Calibri', size=9, color='2a3a5a')

    # Aplicar bordes a todas las celdas de la tabla
    if datos:
        last_row = ROW_HEADER + len(datos)
        for r in range(ROW_HEADER, last_row + 1):
            ws.cell(row=r, column=1).border = Border(
                left=Side(style='medium', color=C_NAVY),
                bottom=ws.cell(row=r, column=1).border.bottom,
                top=ws.cell(row=r, column=1).border.top,
                right=ws.cell(row=r, column=1).border.right,
            )
            ws.cell(row=r, column=num_cols).border = Border(
                right=Side(style='medium', color=C_NAVY),
                bottom=ws.cell(row=r, column=num_cols).border.bottom,
                top=ws.cell(row=r, column=num_cols).border.top,
                left=ws.cell(row=r, column=num_cols).border.left,
            )
        # Línea dorada al final
        final_row = last_row + 1
        ws.row_dimensions[final_row].height = 3
        for c in range(1, num_cols + 1):
            ws.cell(row=final_row, column=c).fill = PatternFill('solid', fgColor=C_GOLD)

        # Pie de tabla
        footer_row = final_row + 1
        ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=num_cols)
        foot = ws.cell(row=footer_row, column=1)
        foot.value = 'Don Bosco La Tola · Quito, Ecuador · donboscolatola.edu.ec · Sistema ClubGest'
        foot.font  = Font(name='Calibri', size=8, color=C_MUTED, italic=True)
        foot.fill  = PatternFill('solid', fgColor=C_NAVY)
        foot.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[footer_row].height = 16

    # Fijar las filas de encabezado para que siempre se vean al desplazarse
    ws.freeze_panes = f'A{ROW_HEADER + 1}'

    # Aplicar filtros a los encabezados para que el usuario pueda buscar
    if datos:
        ws.auto_filter.ref = f'A{ROW_HEADER}:{get_column_letter(num_cols)}{ROW_HEADER + len(datos)}'

    # Crear una segunda hoja para mostrar resumen y estadísticas
    ws2 = wb.create_sheet('Resumen')
    ws2.sheet_view.showGridLines = False

    # Header hoja 2
    ws2.merge_cells('A1:F1')
    t = ws2.cell(row=1, column=1, value='RESUMEN DE INSCRIPCIONES')
    t.font  = Font(name='Calibri', size=14, bold=True, color=C_GOLD2)
    t.fill  = PatternFill('solid', fgColor=C_NAVY)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 28

    if datos:
        # Conteo por club
        clubs_count = {}
        for r in datos:
            c_ = r.get('nombre_club', '—')
            clubs_count[c_] = clubs_count.get(c_, 0) + 1

        # Conteo por nivel
        nivel_count = {}
        for r in datos:
            n_ = r.get('nombre_nivel', '—')
            nivel_count[n_] = nivel_count.get(n_, 0) + 1

        ws2.row_dimensions[2].height = 10

        # Por Club
        ws2.cell(row=3, column=1, value='DISTRIBUCIÓN POR CLUB').font = Font(name='Calibri', size=10, bold=True, color=C_NAVY)
        ws2.row_dimensions[3].height = 18
        for ci, h in enumerate(['Club', 'Inscritos', '% del total'], 1):
            c = ws2.cell(row=4, column=ci, value=h)
            c.font  = Font(name='Calibri', size=9, bold=True, color=C_GOLD2)
            c.fill  = PatternFill('solid', fgColor=C_BLUE)
            c.alignment = Alignment(horizontal='center', vertical='center')
        ws2.row_dimensions[4].height = 18

        for i, (club, cnt) in enumerate(sorted(clubs_count.items(), key=lambda x: -x[1]), 1):
            r_ = 4 + i
            pct = round(cnt / len(datos) * 100, 1)
            vals = [club, cnt, f'{pct}%']
            bg = C_LIGHT1 if i % 2 == 0 else C_LIGHT2
            for ci, v in enumerate(vals, 1):
                c = ws2.cell(row=r_, column=ci, value=v)
                c.fill = PatternFill('solid', fgColor=bg)
                c.font = Font(name='Calibri', size=9,
                               bold=(ci == 1), color=C_NAVY if ci == 1 else '2a3a5a')
                c.alignment = Alignment(horizontal='left' if ci == 1 else 'center',
                                         vertical='center', indent=1 if ci == 1 else 0)

        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 14
        ws2.column_dimensions['C'].width = 14

        # Gráfico de barras
        try:
            chart = BarChart()
            chart.type     = 'col'
            chart.title    = 'Inscritos por Club'
            chart.y_axis.title = 'Cantidad'
            chart.x_axis.title = 'Club'
            chart.style    = 10
            chart.width    = 18
            chart.height   = 12

            data_ref  = Reference(ws2, min_col=2, min_row=4, max_row=4 + len(clubs_count))
            cats_ref  = Reference(ws2, min_col=1, min_row=5, max_row=4 + len(clubs_count))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            ws2.add_chart(chart, 'E3')
        except Exception:
            pass

    # Configurar opciones de impresión del documento
    ws.page_setup.orientation    = 'landscape'
    ws.page_setup.paperSize      = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage      = True
    ws.page_setup.fitToWidth     = 1
    ws.page_setup.fitToHeight    = 0
    ws.print_title_rows = f'{ROW_HEADER}:{ROW_HEADER}'

    # Guardar el archivo Excel en memoria y devolverlo como descarga
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'informe_{tipo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )